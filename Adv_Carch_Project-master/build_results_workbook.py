#!/usr/bin/env python3
"""
Build an Excel workbook from the experiment result text files.

Output:
  - All_Raw / All_Average: workbook-wide summary sheets
  - <cache>_Raw / <cache>_Average: per-cache sheets for each design
Repeated metadata columns are visually grouped by blanking duplicate values in
contiguous runs.
"""

from __future__ import annotations

import ast
import configparser
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RESULTS_DIR = ROOT / "results"
OUTPUT_XLSX = RESULTS_DIR / "attack_summary.xlsx"

CACHE_DIR_MAP = {
    "mirage": ROOT / "Mirage_cache_occupancy",
    "ceaser": ROOT / "Ceaser_cache_occupancy",
    "ceaser_s": ROOT / "Ceaser-s_cache_occupancy",
    "scatter": ROOT / "ScatterCache_cache_occupancy",
    "normal": ROOT / "Normal_cache_occupancy",
}


FILENAME_RE = re.compile(
    r"^outfile_v1_bit_(?P<bit>[01])_(?P<architecture_mode>multibank|hybrid2|region4)_"
    r"banks(?P<num_banks>\d+)_banks_(?P<target_banks>[0-9]+(?:-[0-9]+)*)_"
    r"regions_(?P<target_regions>[0-9]+(?:-[0-9]+)*)"
    r"\.txt$"
)


def parse_config(cache_name: str) -> Dict[str, object]:
    cache_dir = CACHE_DIR_MAP[cache_name]
    cfg_path = cache_dir / "config.ini"
    parser = configparser.ConfigParser()
    parser.read(cfg_path)
    if not parser.sections():
        raise FileNotFoundError(f"Could not read config file: {cfg_path}")
    section = parser[parser.sections()[0]]

    out: Dict[str, object] = {
        "cache_size_bytes": int(section.get("cache-size", 0)),
        "num_blocks_per_set": int(section.get("num-blocks-per-set", 0)),
        "num_additional_tags": int(section.get("num-additional-tags", 0)),
        "num_partitions": int(section.get("num-partitions", 0)),
        "num_words_per_block": int(section.get("num-words-per-block", 0)),
        "num_addr_bits": int(section.get("num-addr-bits", 0)),
        "replacement_policy": section.get("replacement-policy", ""),
    }
    return out


def parse_filename(path: Path) -> Dict[str, object]:
    m = FILENAME_RE.match(path.name)
    if not m:
        raise ValueError(f"Unrecognized result filename: {path.name}")
    info = m.groupdict()
    return {
        "bit": int(info["bit"]),
        "ratio": "0.75" if info["architecture_mode"] == "hybrid2" else "",
        "attack_mode": info["architecture_mode"],
        "architecture_mode": info["architecture_mode"],
        "num_banks": int(info["num_banks"]),
        "target_banks": info["target_banks"].replace("-", ","),
        "target_regions": info["target_regions"].replace("-", ","),
        "banks_to_attack": len(info["target_banks"].split("-")),
    }


def parse_rows(path: Path) -> Tuple[str, List[List[int]], int]:
    rows: List[List[int]] = []
    max_miss_cols = 0
    fmt = "single"
    with path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row = ast.literal_eval(line)
            if not isinstance(row, (list, tuple)) or len(row) < 3:
                continue
            row_list = list(row)
            rows.append(row_list)
            miss_cols = len(row_list) - 2
            max_miss_cols = max(max_miss_cols, miss_cols)
            if miss_cols == 2:
                fmt = "dual"
            elif miss_cols > 2:
                fmt = "multi"
    return fmt, rows, max_miss_cols


def miss_column_names(fmt: str, count: int) -> List[str]:
    if count <= 0:
        return []
    if count == 1:
        return ["misses"]
    if fmt == "dual" and count == 2:
        return ["region0_misses", "region1_misses"]
    if fmt == "multi":
        return [f"bank{i}_misses" for i in range(count)]
    return [f"misses_{i+1}" for i in range(count)]


def build_tables() -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame], Dict[str, pd.DataFrame]]:
    raw_rows: List[Dict[str, object]] = []
    avg_rows: List[Dict[str, object]] = []
    raw_by_cache: Dict[str, List[Dict[str, object]]] = {}
    avg_by_cache: Dict[str, List[Dict[str, object]]] = {}

    result_files = sorted(RESULTS_DIR.glob("*/*/outfile_v1_bit_*.txt"))
    if not result_files:
        raise FileNotFoundError(f"No result files found under {RESULTS_DIR}")

    for path in result_files:
        cache_name = path.parent.name
        case_name = path.parent.parent.name
        cfg = parse_config(cache_name)
        file_info = parse_filename(path)
        fmt, rows, max_miss_cols = parse_rows(path)

        # Build raw rows and assign trial numbers per occupancy.
        trial_counters: Dict[int, int] = {}
        by_occ: Dict[int, List[List[int]]] = {}

        for row in rows:
            occ = int(row[0])
            rec_acc = int(row[1])
            trial_counters[occ] = trial_counters.get(occ, 0) + 1
            trial_idx = trial_counters[occ]

            row_dict: Dict[str, object] = {
                "cache": cache_name,
                "case": case_name,
                "bit": file_info["bit"],
                "ratio": file_info["ratio"],
                "attack_mode": file_info["attack_mode"],
                "architecture_mode": file_info["architecture_mode"],
                "num_banks": file_info["num_banks"],
                "banks_to_attack": file_info["banks_to_attack"],
                "target_banks": file_info["target_banks"],
                "target_regions": file_info["target_regions"],
                "file_format": fmt,
                "occupancy_pct": occ,
                "receiver_accesses": rec_acc,
                "trial_idx": trial_idx,
                **cfg,
            }

            misses = row[2:]
            total_misses = 0
            raw_miss_cols = miss_column_names(fmt, max_miss_cols)
            for i, col_name in enumerate(raw_miss_cols):
                val = int(misses[i]) if i < len(misses) else None
                row_dict[col_name] = val
                if val is not None:
                    total_misses += val
            row_dict["total_misses"] = total_misses if raw_miss_cols else None
            raw_rows.append(row_dict)
            raw_by_cache.setdefault(cache_name, []).append(row_dict)

            by_occ.setdefault(occ, []).append(row)

        # Build averaged rows per occupancy.
        for occ, occ_rows in sorted(by_occ.items()):
            receiver_accesses = int(occ_rows[0][1])
            avg_dict: Dict[str, object] = {
                "cache": cache_name,
                "case": case_name,
                "bit": file_info["bit"],
                "ratio": file_info["ratio"],
                "attack_mode": file_info["attack_mode"],
                "architecture_mode": file_info["architecture_mode"],
                "num_banks": file_info["num_banks"],
                "banks_to_attack": file_info["banks_to_attack"],
                "target_banks": file_info["target_banks"],
                "target_regions": file_info["target_regions"],
                "file_format": fmt,
                "occupancy_pct": occ,
                "receiver_accesses": receiver_accesses,
                "num_trials": len(occ_rows),
                **cfg,
            }

            miss_lists = list(zip(*[r[2:] for r in occ_rows]))
            total_values = [sum(r[2:]) for r in occ_rows]
            avg_miss_cols = miss_column_names(fmt, max_miss_cols)
            for i, col_name in enumerate(avg_miss_cols):
                vals = [int(v) for v in miss_lists[i]] if i < len(miss_lists) else []
                avg_dict[f"mean_{col_name}"] = sum(vals) / len(vals) if vals else None
            avg_dict["mean_total_misses"] = sum(total_values) / len(total_values) if total_values else None
            avg_rows.append(avg_dict)
            avg_by_cache.setdefault(cache_name, []).append(avg_dict)

    raw_df = pd.DataFrame(raw_rows)
    avg_df = pd.DataFrame(avg_rows)

    raw_df = raw_df.sort_values(
        ["cache", "bit", "ratio", "attack_mode", "num_banks", "banks_to_attack", "occupancy_pct", "trial_idx"],
        kind="stable",
    ).reset_index(drop=True)
    avg_df = avg_df.sort_values(
        ["cache", "bit", "ratio", "attack_mode", "num_banks", "banks_to_attack", "occupancy_pct"],
        kind="stable",
    ).reset_index(drop=True)

    raw_cache_dfs = {
        cache_name: pd.DataFrame(rows).sort_values(
            ["bit", "ratio", "attack_mode", "num_banks", "banks_to_attack", "occupancy_pct", "trial_idx"],
            kind="stable"
        ).reset_index(drop=True)
        for cache_name, rows in raw_by_cache.items()
    }
    avg_cache_dfs = {
        cache_name: pd.DataFrame(rows).sort_values(
            ["bit", "ratio", "attack_mode", "num_banks", "banks_to_attack", "occupancy_pct"],
            kind="stable"
        ).reset_index(drop=True)
        for cache_name, rows in avg_by_cache.items()
    }

    return raw_df, avg_df, raw_cache_dfs, avg_cache_dfs


def sparsify_repeated_values(df: pd.DataFrame, group_cols: List[str]) -> pd.DataFrame:
    """Blank repeated values in contiguous runs for selected columns."""
    if df.empty:
        return df
    out = df.copy()
    for col in group_cols:
        if col not in out.columns:
            continue
        out[col] = out[col].astype(object)
        prev = out[col].shift(1)
        out.loc[out[col].eq(prev), col] = ""
    return out


def set_column_widths(ws) -> None:
    """Set widths so headers and values remain visible."""
    for idx, column_cells in enumerate(ws.iter_cols(), start=1):
        header = column_cells[0].value
        max_len = len(str(header)) if header is not None else 0
        sample_limit = min(len(column_cells), 250)
        for cell in column_cells[1:sample_limit]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 28)


def outline_experiment_blocks(ws, df: pd.DataFrame, block_cols: List[str]) -> None:
    """Collapse each contiguous experiment block under its first row."""
    if df.empty or ws.max_row <= 2:
        return

    # Keep the summary row above the grouped detail rows.
    ws.sheet_properties.outlinePr.summaryBelow = False

    keys = df[block_cols].astype(str).agg("|".join, axis=1).tolist()
    start = 0
    for i in range(1, len(keys) + 1):
        if i == len(keys) or keys[i] != keys[start]:
            # Excel row numbers are 1-based, with row 1 as the header.
            first_data_row = start + 2
            last_data_row = i + 1
            if last_data_row > first_data_row:
                ws.row_dimensions.group(
                    first_data_row + 1,
                    last_data_row,
                    outline_level=1,
                    hidden=True,
                )
            start = i


def main() -> None:
    raw_df, avg_df, raw_cache_dfs, avg_cache_dfs = build_tables()

    # Preserve unsparsified copies for outlining; sparsify only the written tables.
    raw_outline_df = raw_df.copy()
    avg_outline_df = avg_df.copy()
    raw_cache_outline_dfs = {k: v.copy() for k, v in raw_cache_dfs.items()}
    avg_cache_outline_dfs = {k: v.copy() for k, v in avg_cache_dfs.items()}

    raw_group_cols = [
        "cache",
        "bit",
        "ratio",
        "attack_mode",
        "num_banks",
        "banks_to_attack",
        "file_format",
        "occupancy_pct",
        "receiver_accesses",
        "cache_size_bytes",
        "num_blocks_per_set",
        "num_additional_tags",
        "num_partitions",
        "num_words_per_block",
        "num_addr_bits",
        "replacement_policy",
    ]
    avg_group_cols = [
        "cache",
        "bit",
        "ratio",
        "attack_mode",
        "num_banks",
        "banks_to_attack",
        "file_format",
        "occupancy_pct",
        "receiver_accesses",
        "cache_size_bytes",
        "num_blocks_per_set",
        "num_additional_tags",
        "num_partitions",
        "num_words_per_block",
        "num_addr_bits",
        "replacement_policy",
    ]

    raw_df = sparsify_repeated_values(raw_df, raw_group_cols)
    avg_df = sparsify_repeated_values(avg_df, avg_group_cols)
    raw_cache_dfs = {k: sparsify_repeated_values(v, raw_group_cols) for k, v in raw_cache_dfs.items()}
    avg_cache_dfs = {k: sparsify_repeated_values(v, avg_group_cols) for k, v in avg_cache_dfs.items()}

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        raw_df.to_excel(writer, index=False, sheet_name="All_Raw")
        avg_df.to_excel(writer, index=False, sheet_name="All_Average")

        for cache_name in sorted(raw_cache_dfs):
            raw_cache_dfs[cache_name].to_excel(
                writer, index=False, sheet_name=f"{cache_name}_Raw"
            )
            avg_cache_dfs[cache_name].to_excel(
                writer, index=False, sheet_name=f"{cache_name}_Avg"
            )

        sheet_to_df = {
            "All_Raw": raw_outline_df,
            "All_Average": avg_outline_df,
        }
        for cache_name in raw_cache_outline_dfs:
            sheet_to_df[f"{cache_name}_Raw"] = raw_cache_outline_dfs[cache_name]
            sheet_to_df[f"{cache_name}_Avg"] = avg_cache_outline_dfs[cache_name]

        block_cols = ["cache", "bit", "ratio", "attack_mode", "num_banks", "banks_to_attack", "file_format"]

        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            set_column_widths(ws)
            outline_experiment_blocks(ws, sheet_to_df[ws.title], block_cols)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Raw rows: {len(raw_df)}")
    print(f"Average rows: {len(avg_df)}")


if __name__ == "__main__":
    main()
